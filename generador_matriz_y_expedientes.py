import os
import fitz  # PyMuPDF
import re
import pandas as pd
import io
from PIL import Image
import pytesseract
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from copy import copy

# 📍 Obtenemos la ruta exacta de la carpeta donde está tu programa
RUTA_PROYECTO = os.path.dirname(os.path.abspath(__file__))

# ⚙️ CONFIGURACIÓN AUTOMÁTICA CORES-PLATAFORMA PARA TESSERACT
if os.name == 'nt':
    ruta_windows = r'C:\Teseract\tesseract.exe'
    if os.path.exists(ruta_windows):
        pytesseract.pytesseract.tesseract_cmd = ruta_windows
    
    ruta_tessdata_local = os.path.join(RUTA_PROYECTO, 'tessdata')
    os.environ['TESSDATA_PREFIX'] = ruta_tessdata_local
else:
    ruta_mac = '/opt/homebrew/bin/tesseract'
    if os.path.exists(ruta_mac):
        pytesseract.pytesseract.tesseract_cmd = ruta_mac

def limpiar_nan(valor):
    if pd.isna(valor) or str(valor).strip().lower() in ['nan', 'nat', 'none', 'null']: return ""
    return str(valor).strip()

def limpiar_numero_formato(valor):
    texto = limpiar_nan(valor)
    return texto[:-2] if texto.endswith('.0') else texto

def obtener_valor_columna(df_row, posibles_nombres):
    for col in posibles_nombres:
        if col in df_row.index:
            return df_row[col]
    return ""

def formatear_documento_identidad(valor):
    texto = limpiar_numero_formato(valor)
    if texto.isdigit() and len(texto) < 10:
        return texto.zfill(10)
    return texto.upper()

def capturar_estilo_celda(celda):
    return {
        "font": copy(celda.font),
        "alignment": copy(celda.alignment),
        "border": copy(celda.border),
        "fill": copy(celda.fill),
        "number_format": celda.number_format,
    }

def aplicar_estilo_celda(celda, estilo):
    celda.font = estilo["font"]
    celda.alignment = estilo["alignment"]
    celda.border = estilo["border"]
    celda.fill = estilo["fill"]
    celda.number_format = estilo["number_format"]

def generar_sistema_y_matriz(ruta_pdf_maestro, ruta_excel_general, ruta_salida_grupo, crear_carpetas=True, exportar_lote_dinardap=False, numero_grupo="", nombre_encabezado="", callback=None):
    
    def notificar(mensaje=None, progreso=None, texto_progreso=None):
        if mensaje: print(mensaje)
        if callback: callback(mensaje, progreso, texto_progreso)

    notificar("🚀 Iniciando sistema de Consolidación Inteligente (V18 - Calibración Fiel Plantilla)...", 2, "Preparando...")

    ruta_plantilla = os.path.join(RUTA_PROYECTO, "PLANTILLA.xlsx")

    if not os.path.exists(ruta_plantilla):
        notificar("❌ ERROR: No se encontró 'PLANTILLA.xlsx' en la carpeta.", 0, "Error")
        raise FileNotFoundError("Plantilla no encontrada.")

    dict_bancos = {
        'BANCO BOLIVARIANO': '1007', 'BANCO DE GUAYAQUIL': '1006', 'BANCO DE LOJA': '1025',
        'BANCO DEL AUSTRO': '1004', 'BANCO DEL PACIFICO': '1028', 'BANCO DEL PICHINCHA': '1029',
        'BANCO PRODUBANCO': '1033', 'COOPERATIVA DE AHORRO Y CREDITO 29 DE OCTUBRE LTDA.': '1122',
        'COOPERATIVA DE AHORRO Y CREDITO VICENTINA MANUEL ESTEBAN GODOY ORTEGA LTDA.': '2129'
    }
    
    ruta_bancos_xlsx = os.path.join(RUTA_PROYECTO, "bancos.xlsx")
    try:
        if os.path.exists(ruta_bancos_xlsx):
            df_bancos = pd.read_excel(ruta_bancos_xlsx)
            for _, row in df_bancos.iterrows():
                nombre_b = str(row.get('Nombre', '')).strip().upper()
                codigo_b = str(row.get('CODIGO', '')).strip()
                if nombre_b and nombre_b != 'NAN': 
                    dict_bancos[nombre_b] = codigo_b
            notificar("🏦 Base de datos 'bancos.xlsx' cargada con éxito.", 3, "Bancos listos")
    except Exception:
        notificar("⚠️ Error al cargar bancos.xlsx, usando respaldo interno.", 3)

    try:
        hojas = pd.read_excel(ruta_excel_general, sheet_name=None, dtype=str)
        for nombre_hoja, df_hoja in hojas.items():
            df_hoja['HOJA_ORIGEN'] = nombre_hoja.strip().upper()
        df_general = pd.concat(hojas.values(), ignore_index=True)
        df_general.columns = df_general.columns.str.strip().str.upper()
        df_general['CEDULA_STR'] = df_general['CEDULA'].apply(formatear_documento_identidad)
        notificar(f"📊 Base de datos Excel cargada ({len(df_general)} registros).", 5, "Cargando Excel...")
    except Exception as e:
        notificar(f"❌ Error crítico al leer el Excel General: {e}", 0)
        raise e

    if not os.path.exists(ruta_salida_grupo):
        os.makedirs(ruta_salida_grupo)

    registros_matriz_final = []
    paginas_fallidas = [] 
    
    doc_maestro = fitz.open(ruta_pdf_maestro)
    total_paginas = len(doc_maestro)
    
    notificar("🔎 Escaneando contratos con Inteligencia Artificial...", 10, f"0/{total_paginas}")
    
    for num_pagina in range(total_paginas):
        numero_real_pagina = num_pagina + 1  
        
        progreso_actual = 10 + int((numero_real_pagina / total_paginas) * 75)
        texto_contador = f"{numero_real_pagina}/{total_paginas}"
        notificar(None, progreso_actual, texto_contador)
        
        pagina = doc_maestro[num_pagina]
        pix = pagina.get_pixmap(matrix=fitz.Matrix(3, 3))  
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        texto = pytesseract.image_to_string(img, lang='spa').replace('\n', ' ')
        
        texto_corregido = texto.replace("8EA", "BEA").replace("8UL", "VUL").replace("0IS", "DIS").upper()
        
        if "CONTRATO" not in texto_corregido and "DBU-" not in texto_corregido:
            paginas_fallidas.append(numero_real_pagina)
            registros_matriz_final.append({
                'NUM_CONTRATO': f"⚠️ PÁG {numero_real_pagina} FALLIDA",
                'PERIODO': 'LLENAR MANUAL', 'FACULTAD': 'LLENAR MANUAL', 'CARRERA': 'LLENAR MANUAL',
                'TIPO_DOCUMENTO': '', 'CEDULA_O_PASAPORTE': 'ILEGIBLE',
                'NOMBRE': "❌ RAZÓN: El OCR no reconoció la palabra CONTRATO (Escaneo muy borroso)",
                'PROMEDIO': '', 'CORREO ELECTRONICO': '', 'TELEFONO ': '', 'DIRECCION': '',
                'NUM_CUENTA': '', 'TIPO_CUENTA': '', 'ENTIDAD_BANCARIA': '', 'CODIGO BANCO': '', 'VALOR': '',
                'TIPO_BECA': '', 'NUMERO_GRUPO': numero_grupo, 'PAGINA': numero_real_pagina
            })
            continue 

        matches_prefijo = re.findall(r"\d{4}\s*-\s*\d{4}", texto_corregido)
        prefijo_contrato = matches_prefijo[0].replace(" ", "") if matches_prefijo else ""

        matches_sufijo = re.findall(r"(?:VUL|BEA|DIS)\s*-\s*\d+", texto_corregido)
        sufijo_contrato = matches_sufijo[-1].replace(" ", "") if matches_sufijo else ""

        if prefijo_contrato and sufijo_contrato: contrato_final = f"{prefijo_contrato}-{sufijo_contrato}"
        elif sufijo_contrato: contrato_final = sufijo_contrato
        elif prefijo_contrato: contrato_final = prefijo_contrato
        else: contrato_final = "NO_DETECTADO"

        if "BECA POR DISCAPACIDAD" in texto_corregido: tipo_beca = "DISCAPACIDAD"
        elif "BECA POR EXCELENCIA" in texto_corregido: tipo_beca = "EXCELENCIA"
        elif "BECA POR VULNERABILIDAD" in texto_corregido: tipo_beca = "VULNERABILIDAD"
        elif "BEA" in contrato_final: tipo_beca = "EXCELENCIA"
        elif "VUL" in contrato_final: tipo_beca = "VULNERABILIDAD"
        elif "DIS" in contrato_final: tipo_beca = "DISCAPACIDAD"
        else: tipo_beca = "OTRA"

        match_periodo = re.search(r"PERIODO ACAD[EÉ]MICO\s+(.*?)\s+CONTRATO", texto_corregido)
        periodo_pdf = match_periodo.group(1).strip() if match_periodo else ""
        periodo_pdf = re.sub(r"[\u2012\u2013\u2014\u2015]", "-", periodo_pdf)

        cedula_extraida = None
        texto_limpio_busqueda = re.sub(r'[\s\-]', '', texto_corregido) 
        
        patron_zona = r"(?:C[EÉ]DULA|PASAPORTE|C\.C\.)(?:NRO|NO)?([A-Z0-9]{7,20})"
        match_zona = re.search(patron_zona, texto_limpio_busqueda)
        
        if match_zona:
            zona_sucia = match_zona.group(1)
            for doc_excel in df_general['CEDULA_STR'].unique():
                if len(doc_excel) >= 7 and doc_excel in zona_sucia:
                    cedula_extraida = doc_excel
                    break
        
        if not cedula_extraida:
            for doc_excel in df_general['CEDULA_STR'].unique():
                if len(doc_excel) >= 7 and doc_excel in texto_limpio_busqueda:
                    cedula_extraida = doc_excel
                    break

        if not cedula_extraida:
            paginas_fallidas.append(numero_real_pagina)
            registros_matriz_final.append({
                'NUM_CONTRATO': f"⚠️ PÁG {numero_real_pagina} FALLIDA",
                'PERIODO': 'LLENAR MANUAL', 'FACULTAD': 'LLENAR MANUAL', 'CARRERA': 'LLENAR MANUAL',
                'TIPO_DOCUMENTO': '', 'CEDULA_O_PASAPORTE': 'NO DETECTADA',
                'NOMBRE': "❌ RAZÓN: No se detectó ningún documento de identidad válido en esta página",
                'PROMEDIO': '', 'CORREO ELECTRONICO': '', 'TELEFONO ': '', 'DIRECCION': '',
                'NUM_CUENTA': '', 'TIPO_CUENTA': '', 'ENTIDAD_BANCARIA': '', 'CODIGO BANCO': '', 'VALOR': '',
                'TIPO_BECA': '', 'NUMERO_GRUPO': numero_grupo, 'PAGINA': numero_real_pagina
            })
            continue

        match_excel = df_general[df_general['CEDULA_STR'] == cedula_extraida]
        
        if match_excel.empty:
            paginas_fallidas.append(numero_real_pagina)
            registros_matriz_final.append({
                'NUM_CONTRATO': f"⚠️ PÁG {numero_real_pagina} FALLIDA",
                'PERIODO': 'LLENAR MANUAL', 'FACULTAD': 'LLENAR MANUAL', 'CARRERA': 'LLENAR MANUAL',
                'TIPO_DOCUMENTO': '', 'CEDULA_O_PASAPORTE': cedula_extraida,
                'NOMBRE': f"❌ RAZÓN: El documento {cedula_extraida} NO EXISTE en el Excel base",
                'PROMEDIO': '', 'CORREO ELECTRONICO': '', 'TELEFONO ': '', 'DIRECCION': '',
                'NUM_CUENTA': '', 'TIPO_CUENTA': '', 'ENTIDAD_BANCARIA': '', 'CODIGO BANCO': '', 'VALOR': '',
                'TIPO_BECA': '', 'NUMERO_GRUPO': numero_grupo, 'PAGINA': numero_real_pagina
            })
            continue
            
        datos_estudiante = match_excel.iloc[0]
        nombre_oficial = str(datos_estudiante['NOMBRE']).strip()

        if tipo_beca == "OTRA":
            hoja_origen = limpiar_nan(obtener_valor_columna(datos_estudiante, ['HOJA_ORIGEN'])).upper()
            if "DISCAPACIDAD" in hoja_origen: tipo_beca = "DISCAPACIDAD"
            elif "EXCELENCIA" in hoja_origen: tipo_beca = "EXCELENCIA"
            elif "VULNERABILIDAD" in hoja_origen: tipo_beca = "VULNERABILIDAD"
        
        if crear_carpetas:
            nombre_carpeta = f"{numero_real_pagina}_{cedula_extraida}_{nombre_oficial.replace(' ', '_')}"
            ruta_nueva_carpeta = os.path.join(ruta_salida_grupo, nombre_carpeta)
            os.makedirs(ruta_nueva_carpeta, exist_ok=True)
            
            ruta_destino_pdf = os.path.join(ruta_nueva_carpeta, f"C_{cedula_extraida}.pdf")
            nuevo_pdf = fitz.open()
            nuevo_pdf.insert_pdf(doc_maestro, from_page=num_pagina, to_page=num_pagina)
            nuevo_pdf.save(ruta_destino_pdf)
            nuevo_pdf.close()
        
        periodo_excel = limpiar_nan(obtener_valor_columna(datos_estudiante, ['* PERIODO', 'PERIODO', 'PERÍODO']))
        periodo_val = periodo_pdf if periodo_pdf else periodo_excel

        contrato_excel = limpiar_nan(obtener_valor_columna(datos_estudiante, [
            'NUMERO DE CONTRATO', 'NÚMERO DE CONTRATO', 'NO CONTRATO', 'N° CONTRATO', 'NUM CONTRATO', 'NUMERO CONTRATO', 'N CONTRATO'
        ]))
        contrato_val = contrato_final if contrato_final != "NO_DETECTADO" else (contrato_excel or contrato_final)

        facultad_val = limpiar_nan(obtener_valor_columna(datos_estudiante, ['FACULTAD']))
        carrera_val = limpiar_nan(obtener_valor_columna(datos_estudiante, ['CARRERA']))
        correo_val = limpiar_nan(obtener_valor_columna(datos_estudiante, ['CORREO', 'CORREO INSTITUCIONAL', 'EMAIL']))
        
        promedio_raw = limpiar_nan(obtener_valor_columna(datos_estudiante, ['GENERAL', 'PROMEDIO']))
        promedio_final = ""
        if tipo_beca == "EXCELENCIA" and promedio_raw:
            try:
                prom_str = str(promedio_raw).replace(",", ".")
                promedio_final = float(prom_str)
            except ValueError:
                promedio_final = promedio_raw

        tel_raw = limpiar_numero_formato(obtener_valor_columna(datos_estudiante, ['CELULAR', 'TELEFONO', 'TELÉFONO']))
        telefono_final = tel_raw.zfill(10) if tel_raw else "022542160"

        dir_raw = limpiar_nan(obtener_valor_columna(datos_estudiante, ['DIRECCION', 'DIRECCIÓN']))
        direccion_final = dir_raw if dir_raw else "Ciudadela Universitaria"

        cuenta_final = limpiar_numero_formato(obtener_valor_columna(datos_estudiante, ['NO. CUENTA', 'NO CUENTA', 'CUENTA', 'N° CUENTA']))
        tipo_cuenta_final = limpiar_nan(obtener_valor_columna(datos_estudiante, ['TIPO CUENTA', 'TIPO_CUENTA']))

        match_bloque = re.search(r"CUENTA\s*[:\-]?\s*(AHORROS?|CORRIENTE)?\D*?(\d{7,15})", texto_corregido)
        if match_bloque:
            if not tipo_cuenta_final and match_bloque.group(1):
                tipo_cuenta_final = "AHORROS" if "AHORRO" in match_bloque.group(1) else "CORRIENTE"
            if not cuenta_final and match_bloque.group(2):
                cuenta_final = match_bloque.group(2).strip()

        tipo_cuenta_final = str(tipo_cuenta_final).upper() if tipo_cuenta_final else ""

        banco_raw = limpiar_nan(obtener_valor_columna(datos_estudiante, ['ENTIDAD BANCARIA', 'BANCO'])).upper()
        codigo_banco_final = ""
        banco_encontrado = ""
        bancos_ordenados = sorted(dict_bancos.keys(), key=len, reverse=True)

        if banco_raw:
            for nombre_b in bancos_ordenados:
                if banco_raw in nombre_b or nombre_b in banco_raw:
                    banco_encontrado = nombre_b
                    codigo_banco_final = dict_bancos[nombre_b]
                    break

        if not banco_encontrado:
            for nombre_b in bancos_ordenados:
                if nombre_b in texto_corregido:
                    banco_encontrado = nombre_b
                    codigo_banco_final = dict_bancos[nombre_b]
                    break

        if not banco_encontrado:
            if "PICHINCHA" in texto_corregido: banco_encontrado, codigo_banco_final = "BANCO DEL PICHINCHA", "1029"
            elif "GUAYAQUIL" in texto_corregido: banco_encontrado, codigo_banco_final = "BANCO DE GUAYAQUIL", "1006"
            elif "PACIFICO" in texto_corregido or "PACÍFICO" in texto_corregido: banco_encontrado, codigo_banco_final = "BANCO DEL PACIFICO", "1028"
            elif "AUSTRO" in texto_corregido: banco_encontrado, codigo_banco_final = "BANCO DEL AUSTRO", "1004"
            elif "BOLIVARIANO" in texto_corregido: banco_encontrado, codigo_banco_final = "BANCO BOLIVARIANO", "1007"
            elif "LOJA" in texto_corregido: banco_encontrado, codigo_banco_final = "BANCO DE LOJA", "1025"
            elif "29 DE OCTUBRE" in texto_corregido: banco_encontrado, codigo_banco_final = "COOPERATIVA DE AHORRO Y CREDITO 29 DE OCTUBRE LTDA.", "1122"

        if banco_encontrado: banco_raw = banco_encontrado

        registros_matriz_final.append({
            'NUM_CONTRATO': contrato_val,
            'PERIODO': periodo_val,
            'FACULTAD': facultad_val,
            'CARRERA': carrera_val,
            'TIPO_DOCUMENTO': 'CEDULA' if cedula_extraida.isdigit() and len(cedula_extraida) <= 10 else 'PASAPORTE',
            'CEDULA_O_PASAPORTE': cedula_extraida, 
            'NOMBRE': nombre_oficial,
            'PROMEDIO': promedio_final,
            'CORREO ELECTRONICO': correo_val,
            'TELEFONO ': telefono_final, 
            'DIRECCION': direccion_final, 
            'NUM_CUENTA': cuenta_final,
            'TIPO_CUENTA': tipo_cuenta_final, 
            'ENTIDAD_BANCARIA': banco_raw, 
            'CODIGO BANCO': codigo_banco_final, 
            'VALOR': 400,
            'TIPO_BECA': tipo_beca,
            'NUMERO_GRUPO': numero_grupo,
            'PAGINA': numero_real_pagina
        })
        notificar(f"✅ Procesado: {nombre_oficial} (Pág {numero_real_pagina})", progreso_actual, texto_contador)

    doc_maestro.close()

    # =======================================================
    # 4. CONSTRUCCIÓN DE EXCEL FIEL AL 100% (SIN FILAS EXTRAS)
    # =======================================================
    # 🔥 CALIBRACIÓN MILIMÉTRICA BASADA EN TU IMAGEN REAL
    HEADER_ROW         = 1     # Fila 1: Encabezados nativos de la tabla (No BECARIO, NUM_CONTRATO, etc.)
    FIRST_DATA_ROW     = 2     # Fila 2: Celda del Estudiante número 1
    ORIG_LAST_DATA_ROW = 101   # Fila 101: Celda del Estudiante número 100 en plantilla vacía
    N_COLS             = 17    

    FIELD_ORDER = [
        "NUM_CONTRATO", "PERIODO", "FACULTAD", "CARRERA", "TIPO_DOCUMENTO",
        "CEDULA_O_PASAPORTE", "NOMBRE", "PROMEDIO", "CORREO ELECTRONICO",
        "TELEFONO ", "DIRECCION", "NUM_CUENTA", "TIPO_CUENTA",
        "ENTIDAD_BANCARIA", "CODIGO BANCO", "VALOR"
    ]
    CAMPOS_TEXTO = {'CEDULA_O_PASAPORTE', 'TELEFONO ', 'NUM_CUENTA'}

    texto_final = f"{total_paginas}/{total_paginas}"

    if registros_matriz_final:
        try:
            # Determinamos los metadatos en base al primer contrato válido leído
            tipo_lote = 'OTRA'
            for r in registros_matriz_final:
                t = r.get('TIPO_BECA', 'OTRA')
                if t and t in ['EXCELENCIA', 'VULNERABILIDAD', 'DISCAPACIDAD']:
                    tipo_lote = t
                    break

            registros_validos  = []
            registros_anomalos = []
            for r in registros_matriz_final:
                t = r.get('TIPO_BECA', 'OTRA')
                if t and t != 'OTRA' and t != tipo_lote: registros_anomalos.append(r)
                else: registros_validos.append(r)

            if registros_anomalos:
                lineas = [f"⚠️ SE DETECTARON {len(registros_anomalos)} CONTRATO(S) DISTINTOS AL TIPO DEL LOTE ({tipo_lote}) Y FUERON EXCLUIDOS:"]
                for ra in registros_anomalos:
                    lineas.append(f"   • Pág {ra.get('PAGINA','-')} | {ra.get('NOMBRE','SIN NOMBRE')} | Contrato: {ra.get('NUM_CONTRATO','-')} | Tipo: {ra.get('TIPO_BECA','-')}")
                notificar("\n".join(lineas), 92, texto_final)

            # Generamos el nombre del archivo usando los datos enviados a la interfaz
            periodo_enc = registros_validos[0].get('PERIODO', '').strip() if registros_validos else ""
            grupo_enc   = numero_grupo.strip() if numero_grupo else ""
            
            partes_nombre_archivo = ["Matriz", "Contratos"]
            if grupo_enc: partes_nombre_archivo.append(f"Grupo_{grupo_enc}")
            if tipo_lote and tipo_lote != 'OTRA': partes_nombre_archivo.append(tipo_lote)
            if periodo_enc: partes_nombre_archivo.append(periodo_enc.replace(" ", "_").replace("-", "_"))
            
            # Formateamos caracteres inválidos para rutas de Windows
            nombre_limpio_excel = re.sub(r'[\\/*?:"<>|]', "", "_".join(partes_nombre_archivo)) + ".xlsx"
            archivo_matriz_salida = os.path.join(ruta_salida_grupo, nombre_limpio_excel)

            notificar("📝 Preparando plantilla base...", 93, texto_final)
            wb = openpyxl.load_workbook(ruta_plantilla)
            ws = wb.active

            # Captura de estilos de bandas originales (Filas 2 y 3)
            estilo_impar = { col: capturar_estilo_celda(ws.cell(row=FIRST_DATA_ROW, column=col)) for col in range(1, N_COLS + 1) }
            estilo_par = { col: capturar_estilo_celda(ws.cell(row=FIRST_DATA_ROW + 1, column=col)) for col in range(1, N_COLS + 1) }
            altura_fila_dato = ws.row_dimensions[FIRST_DATA_ROW].height

            # Captura matemática precisa del pie de página original (Fila 102 en adelante)
            pie_filas = []
            for r in range(ORIG_LAST_DATA_ROW + 1, ws.max_row + 1):
                celdas = {}
                for col in range(1, N_COLS + 1):
                    c = ws.cell(row=r, column=col)
                    celdas[col] = {"valor": c.value, "estilo": capturar_estilo_celda(c)}
                pie_filas.append({
                    "offset": r - ORIG_LAST_DATA_ROW,
                    "altura": ws.row_dimensions[r].height,
                    "celdas": celdas,
                })

            # Captura exacta de los merges de los cuadros de firmas
            pie_merges = []
            for mc in list(ws.merged_cells.ranges):
                if mc.min_row > ORIG_LAST_DATA_ROW:
                    pie_merges.append({
                        "offset_min": mc.min_row - ORIG_LAST_DATA_ROW,
                        "offset_max": mc.max_row - ORIG_LAST_DATA_ROW,
                        "min_col": mc.min_col,
                        "max_col": mc.max_col,
                    })

            if "Tabla1" in ws.tables: del ws.tables["Tabla1"]
            
            # Descombinar únicamente el rango de estudiantes para proteger firmas
            for mc in list(ws.merged_cells.ranges):
                if mc.min_row >= FIRST_DATA_ROW:
                    ws.unmerge_cells(str(mc))
            
            # Borrar las 100 filas de prueba antiguas dejando intacta la fila 1 de encabezados
            ws.delete_rows(FIRST_DATA_ROW, ws.max_row - HEADER_ROW)

            notificar("📝 Escribiendo registros estructurados...", 96, texto_final)
            n = len(registros_validos)
            lista_dinardap_rapida = []

            for i, registro in enumerate(registros_validos, start=1):
                row = FIRST_DATA_ROW + i - 1
                estilo_fila = estilo_impar if i % 2 == 1 else estilo_par
                ws.row_dimensions[row].height = altura_fila_dato
                
                ws.cell(row=row, column=1, value=i)
                aplicar_estilo_celda(ws.cell(row=row, column=1), estilo_fila[1])
                
                for j, campo in enumerate(FIELD_ORDER, start=2):
                    c = ws.cell(row=row, column=j, value=registro.get(campo, ""))
                    estilo_aplicar = copy(estilo_fila[j])
                    if campo in CAMPOS_TEXTO: estilo_aplicar["number_format"] = '@'
                    aplicar_estilo_celda(c, estilo_aplicar)

                if (registro.get('CEDULA_O_PASAPORTE', '') not in ['', 'ILEGIBLE', 'NO DETECTADA'] 
                    and 'FALLIDA' not in str(registro.get('NUM_CONTRATO', ''))):
                    lista_dinardap_rapida.append({
                        'CEDULA': registro['CEDULA_O_PASAPORTE'],
                        'NOMBRE': registro['NOMBRE'],
                    })

            ultima_fila_datos = FIRST_DATA_ROW + n - 1

            notificar("📝 Restaurando firmas y pie de página compactos...", 98, texto_final)
            for bloque in pie_filas:
                row = ultima_fila_datos + bloque["offset"]
                if bloque["altura"]:
                    ws.row_dimensions[row].height = bloque["altura"]
                for col, info in bloque["celdas"].items():
                    c = ws.cell(row=row, column=col, value=info["valor"])
                    aplicar_estilo_celda(c, info["estilo"])

            # Recombinar de forma exacta vertical y horizontalmente las firmas
            for m in pie_merges:
                r_min = ultima_fila_datos + m["offset_min"]
                r_max = ultima_fila_datos + m["offset_max"]
                ws.merge_cells(start_row=r_min, start_column=m["min_col"], end_row=r_max, end_column=m["max_col"])

            # Recreación perfecta de la tabla y los filtros desde la Fila 1
            ultima_col = get_column_letter(N_COLS)
            ref = f"A{HEADER_ROW}:{ultima_col}{ultima_fila_datos}"
            tabla = Table(displayName="Tabla1", ref=ref)
            tabla.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight16", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tabla)

            # Contorno de bordes perimetrales
            grueso = Side(style="medium", color="FF000000")
            for col in range(1, N_COLS + 1):
                tc = ws.cell(row=HEADER_ROW, column=col)
                tc.border = Border(left=tc.border.left, right=tc.border.right, top=grueso, bottom=tc.border.bottom)
                bc = ws.cell(row=ultima_fila_datos, column=col)
                bc.border = Border(left=bc.border.left, right=bc.border.right, top=bc.border.top, bottom=grueso)
            for row in range(HEADER_ROW, ultima_fila_datos + 1):
                lc = ws.cell(row=row, column=1)
                lc.border = Border(left=grueso, right=lc.border.right, top=lc.border.top, bottom=lc.border.bottom)
                rc = ws.cell(row=row, column=N_COLS)
                rc.border = Border(left=rc.border.left, right=grueso, top=rc.border.top, bottom=rc.border.bottom)

            wb.save(archivo_matriz_salida)

# Conservada la exportación nativa de Dinardap (Fase 1) CON encabezado
            if exportar_lote_dinardap and lista_dinardap_rapida:
                df_din = pd.DataFrame(lista_dinardap_rapida)
                ruta_din_salida = os.path.join(ruta_salida_grupo, "Lote_Dinardap_Fase1.xlsx")
                # Se cambia header=False a header=True
                df_din[['CEDULA', 'NOMBRE']].to_excel(ruta_din_salida, index=False, header=True)
                notificar("📄 Archivo 'Lote_Dinardap_Fase1.xlsx' exportado con encabezados.")

            resumen = (f"🎉 MATRIZ GENERADA: {n} registros válidos"
                       + (f" | {len(registros_anomalos)} excluidos" if registros_anomalos else "")
                       + (f" | {len(paginas_fallidas)} fallos" if paginas_fallidas else "")
                       + f"\n📄 Archivo guardado como: {os.path.basename(archivo_matriz_salida)}")
            notificar(resumen, 100, texto_final)

        except Exception as e:
            notificar(f"❌ Error al guardar Excel: {e}", 100, "Error")
            raise e
    else:
        notificar("⚠️ No se detectaron registros válidos en el PDF.", 100, texto_final)