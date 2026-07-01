#!/usr/bin/env python3
"""
Generador de listado de becarios a partir de la PLANTILLA.xlsx
================================================================

Toma como base la plantilla original (con su formato, encabezados,
filtros/tabla, fuentes, anchos de columna y configuración de impresión)
y genera un nuevo Excel con tantas filas de datos como registros se
entreguen (1, 100, 500... cualquier cantidad), manteniendo el bloque
de firma ("REVISADO Y ELABORADO POR...") siempre justo debajo del
último registro.

USO
---
1) Con un CSV de datos (recomendado):
   python generar_excel_becas.py datos.csv salida.xlsx

   El CSV debe tener como encabezado EXACTAMENTE estos 16 campos
   (en cualquier orden), separados por coma o punto y coma:
   NUM_CONTRATO, PERIODO, FACULTAD, CARRERA, TIPO_DOCUMENTO,
   CEDULA_O_PASAPORTE, NOMBRE, PROMEDIO, CORREO ELECTRONICO,
   TELEFONO, DIRECCION, NUM_CUENTA, TIPO_CUENTA, ENTIDAD_BANCARIA,
   CODIGO BANCO, VALOR
   (La columna "No  BECARIO" se numera automáticamente, no va en el CSV)

2) Sin argumentos: genera un archivo de ejemplo con datos de muestra,
   útil para probar que el formato se respeta.

Requiere: openpyxl  (pip install openpyxl --break-system-packages)
"""

import sys
import csv
import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

PLANTILLA = Path(__file__).parent / "PLANTILLA.xlsx"
SHEET_NAME = "Hoja1"
HEADER_ROW = 1
FIRST_DATA_ROW = 2
ORIG_LAST_DATA_ROW = 101          # fila de datos #100 en la plantilla original
N_COLS = 17                       # columnas A..Q

# Orden de columnas tal como aparece en la plantilla (B..Q -> índices 2..17)
# A (índice 1) = "No  BECARIO" se autonumera
FIELD_ORDER = [
    "NUM_CONTRATO", "PERIODO", "FACULTAD", "CARRERA", "TIPO_DOCUMENTO",
    "CEDULA_O_PASAPORTE", "NOMBRE", "PROMEDIO", "CORREO ELECTRONICO",
    "TELEFONO", "DIRECCION", "NUM_CUENTA", "TIPO_CUENTA",
    "ENTIDAD_BANCARIA", "CODIGO BANCO", "VALOR",
]


def leer_datos_csv(path):
    """Lee un CSV (coma o punto y coma) y devuelve una lista de dicts."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(f, delimiter=delim)
        # normaliza encabezados (sin espacios extra, sin distinguir mayúsc/minúsc)
        norm_map = {k.strip().upper(): k for k in reader.fieldnames}
        registros = []
        for fila in reader:
            registro = {}
            for campo in FIELD_ORDER:
                clave_real = norm_map.get(campo.strip().upper())
                registro[campo] = fila.get(clave_real, "") if clave_real else ""
            registros.append(registro)
        return registros


def datos_de_ejemplo():
    return [
        {
            "NUM_CONTRATO": "CT-001", "PERIODO": "2026-A", "FACULTAD": "Ingeniería",
            "CARRERA": "Sistemas", "TIPO_DOCUMENTO": "Cédula",
            "CEDULA_O_PASAPORTE": "1712345678", "NOMBRE": "Juan Pérez",
            "PROMEDIO": 9.2, "CORREO ELECTRONICO": "juan.perez@correo.edu",
            "TELEFONO": "0991234567", "DIRECCION": "Av. Siempre Viva 123",
            "NUM_CUENTA": "1234567890", "TIPO_CUENTA": "Ahorros",
            "ENTIDAD_BANCARIA": "Banco Pichincha", "CODIGO BANCO": "10",
            "VALOR": 350.00,
        },
        {
            "NUM_CONTRATO": "CT-002", "PERIODO": "2026-A", "FACULTAD": "Ciencias",
            "CARRERA": "Biología", "TIPO_DOCUMENTO": "Cédula",
            "CEDULA_O_PASAPORTE": "1798765432", "NOMBRE": "María Gómez",
            "PROMEDIO": 8.7, "CORREO ELECTRONICO": "maria.gomez@correo.edu",
            "TELEFONO": "0987654321", "DIRECCION": "Calle Falsa 456",
            "NUM_CUENTA": "0987654321", "TIPO_CUENTA": "Corriente",
            "ENTIDAD_BANCARIA": "Banco Guayaquil", "CODIGO BANCO": "11",
            "VALOR": 280.50,
        },
    ]


def capturar_estilo_celda(celda):
    return {
        "font": copy.copy(celda.font),
        "alignment": copy.copy(celda.alignment),
        "border": copy.copy(celda.border),
        "fill": copy.copy(celda.fill),
        "number_format": celda.number_format,
    }


def aplicar_estilo_celda(celda, estilo):
    celda.font = estilo["font"]
    celda.alignment = estilo["alignment"]
    celda.border = estilo["border"]
    celda.fill = estilo["fill"]
    celda.number_format = estilo["number_format"]


def generar(registros, salida):
    wb = load_workbook(PLANTILLA)
    ws = wb[SHEET_NAME]

    # 1) Capturar estilos de las filas de datos plantilla: la plantilla
    #    alterna banda azul (fila 2) y banda blanca (fila 3) por cada
    #    registro, hay que respetar ese patrón en vez de copiar un único
    #    estilo para todas las filas.
    estilo_par = {
        col: capturar_estilo_celda(ws.cell(row=FIRST_DATA_ROW, column=col))
        for col in range(1, N_COLS + 1)
    }
    estilo_impar = {
        col: capturar_estilo_celda(ws.cell(row=FIRST_DATA_ROW + 1, column=col))
        for col in range(1, N_COLS + 1)
    }
    altura_fila_dato = ws.row_dimensions[FIRST_DATA_ROW].height

    # 2) Capturar bloque de pie de página (firma) tal como está en la plantilla
    pie_filas = []
    for r in range(ORIG_LAST_DATA_ROW + 1, ws.max_row + 1):
        celdas = {}
        for col in range(1, N_COLS + 1):
            c = ws.cell(row=r, column=col)
            celdas[col] = {"valor": c.value, "estilo": capturar_estilo_celda(c)}
        pie_filas.append({
            "offset": r - ORIG_LAST_DATA_ROW,
            "altura": ws.row_dimensions[r].height,
            "celdas": celdas,
        })

    pie_merges = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row > ORIG_LAST_DATA_ROW:
            pie_merges.append({
                "offset": mc.min_row - ORIG_LAST_DATA_ROW,
                "min_col": mc.min_col,
                "max_col": mc.max_col,
            })

    # 3) Quitar tabla y desmerge de celdas del pie (se recrean al final)
    if "Tabla1" in ws.tables:
        del ws.tables["Tabla1"]
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))

    # 4) Borrar todo el contenido bajo el encabezado (datos + pie originales)
    ws.delete_rows(FIRST_DATA_ROW, ws.max_row - HEADER_ROW)

    # 5) Escribir filas de datos (1 a N), manteniendo formato exacto
    n = len(registros)
    for i, registro in enumerate(registros, start=1):
        row = FIRST_DATA_ROW + i - 1
        estilo_fila = estilo_par if i % 2 == 1 else estilo_impar
        ws.row_dimensions[row].height = altura_fila_dato
        ws.cell(row=row, column=1, value=i)
        aplicar_estilo_celda(ws.cell(row=row, column=1), estilo_fila[1])
        for j, campo in enumerate(FIELD_ORDER, start=2):
            c = ws.cell(row=row, column=j, value=registro.get(campo, ""))
            aplicar_estilo_celda(c, estilo_fila[j])

    ultima_fila_datos = FIRST_DATA_ROW + n - 1

    # 6) Volver a escribir el bloque de pie de página justo debajo
    for bloque in pie_filas:
        row = ultima_fila_datos + bloque["offset"]
        if bloque["altura"]:
            ws.row_dimensions[row].height = bloque["altura"]
        for col, info in bloque["celdas"].items():
            c = ws.cell(row=row, column=col, value=info["valor"])
            aplicar_estilo_celda(c, info["estilo"])

    for m in pie_merges:
        row = ultima_fila_datos + m["offset"]
        ws.merge_cells(start_row=row, start_column=m["min_col"],
                        end_row=row, end_column=m["max_col"])

    # 7) Recrear la tabla con filtros (autofiltro) ajustada al nuevo rango
    ultima_col = get_column_letter(N_COLS)
    ref = f"A{HEADER_ROW}:{ultima_col}{ultima_fila_datos}"
    tabla = Table(displayName="Tabla1", ref=ref)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight16", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tabla)

    # 8) Borde exterior grueso alrededor de toda la tabla (encabezado + datos)
    grueso = Side(style="medium", color="FF000000")
    primera_col, ultima_col_idx = 1, N_COLS
    for col in range(primera_col, ultima_col_idx + 1):
        top_c = ws.cell(row=HEADER_ROW, column=col)
        top_c.border = Border(left=top_c.border.left, right=top_c.border.right,
                               top=grueso, bottom=top_c.border.bottom)
        bot_c = ws.cell(row=ultima_fila_datos, column=col)
        bot_c.border = Border(left=bot_c.border.left, right=bot_c.border.right,
                               top=bot_c.border.top, bottom=grueso)
    for row in range(HEADER_ROW, ultima_fila_datos + 1):
        left_c = ws.cell(row=row, column=primera_col)
        left_c.border = Border(left=grueso, right=left_c.border.right,
                                top=left_c.border.top, bottom=left_c.border.bottom)
        right_c = ws.cell(row=row, column=ultima_col_idx)
        right_c.border = Border(left=right_c.border.left, right=grueso,
                                 top=right_c.border.top, bottom=right_c.border.bottom)

    wb.save(salida)
    print(f"Generado: {salida}  ({n} registros)")


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        registros = leer_datos_csv(sys.argv[1])
        salida = sys.argv[2] if len(sys.argv) >= 3 else "Listado_Becarios.xlsx"
    else:
        registros = datos_de_ejemplo()
        salida = "Listado_Becarios_Ejemplo.xlsx"
    generar(registros, salida)